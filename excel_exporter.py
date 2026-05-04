from collections import defaultdict
from decimal import Decimal
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from invoice_extractor import InvoiceRecord


DATE_FMT = "DD/MM/YYYY"
CURRENCY_FMT = 'R$ #,##0.00'


def _auto_size_columns(sheet) -> None:
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _month_year_sort_key(month_key: str) -> Tuple[int, int]:
    month_str, year_str = month_key.split("/")
    return int(year_str), int(month_str)


def export_to_excel(records: List[InvoiceRecord], output_path: str) -> None:
    wb = Workbook()

    # Aba 1: Notas Fiscais
    ws_notas = wb.active
    ws_notas.title = "Notas Fiscais"
    ws_notas.append(["Data", "Fornecedor", "Valor"])

    for cell in ws_notas[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sorted_records = sorted(records, key=lambda r: r.data)
    for record in sorted_records:
        ws_notas.append([record.data, record.fornecedor, float(record.valor)])

    for row in ws_notas.iter_rows(min_row=2, max_col=3):
        row[0].number_format = DATE_FMT
        row[2].number_format = CURRENCY_FMT

    _auto_size_columns(ws_notas)

    # Aba 2: Consolidação
    ws_cons = wb.create_sheet("Consolidação")
    ws_cons.append(["Fornecedor", "Mês/Ano", "Total Gasto", "Quantidade de Notas"])

    for cell in ws_cons[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    grouped_by_supplier_month = defaultdict(lambda: {"total": Decimal("0"), "qtd": 0})
    grouped_by_supplier_total = defaultdict(lambda: {"total": Decimal("0"), "qtd": 0})

    for record in sorted_records:
        month_key = record.data.strftime("%m/%Y")

        supplier_month_key = (record.fornecedor, month_key)
        grouped_by_supplier_month[supplier_month_key]["total"] += record.valor
        grouped_by_supplier_month[supplier_month_key]["qtd"] += 1

        grouped_by_supplier_total[record.fornecedor]["total"] += record.valor
        grouped_by_supplier_total[record.fornecedor]["qtd"] += 1

    # Tabela 1: Fornecedor + Mês/Ano (ordenada por Mês/Ano crescente)
    sorted_supplier_month_rows = sorted(
        grouped_by_supplier_month.items(),
        key=lambda item: (
            *_month_year_sort_key(item[0][1]),
            item[0][0].lower(),
        ),
    )

    first_table_start_row = 2
    for (fornecedor, month_key), agg in sorted_supplier_month_rows:
        ws_cons.append([fornecedor, month_key, float(agg["total"]), agg["qtd"]])

    first_table_end_row = ws_cons.max_row

    for row in ws_cons.iter_rows(min_row=first_table_start_row, max_row=first_table_end_row, max_col=4):
        row[1].alignment = Alignment(horizontal="center")
        row[2].number_format = CURRENCY_FMT
        row[3].alignment = Alignment(horizontal="center")

    # Linha em branco entre as tabelas
    ws_cons.append([])

    # Tabela 2: Total por fornecedor em todos os meses
    second_table_header_row = ws_cons.max_row + 1
    ws_cons.append(["Fornecedor", "Total Geral (Todos os Meses)", "Quantidade de Notas"])

    for cell in ws_cons[second_table_header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sorted_supplier_total_rows = sorted(
        grouped_by_supplier_total.items(),
        key=lambda item: (-item[1]["total"], item[0].lower()),
    )

    second_table_start_row = ws_cons.max_row + 1
    for fornecedor, agg in sorted_supplier_total_rows:
        ws_cons.append([fornecedor, float(agg["total"]), agg["qtd"]])

    second_table_end_row = ws_cons.max_row

    for row in ws_cons.iter_rows(min_row=second_table_start_row, max_row=second_table_end_row, max_col=3):
        row[1].number_format = CURRENCY_FMT
        row[2].alignment = Alignment(horizontal="center")

    _auto_size_columns(ws_cons)

    wb.save(output_path)
